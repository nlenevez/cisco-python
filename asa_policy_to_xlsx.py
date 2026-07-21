#!/usr/bin/env python3
"""asa_policy_to_xlsx.py -- turn a Cisco ASA running-config into a readable XLSX.

Parses network/service objects, object-groups (network/service/protocol/icmp-type),
extended access-lists and their interface bindings (access-group), resolves the
object/group references, and writes a multi-sheet Excel workbook:

  * Summary            -- counts + ACL-to-interface bindings
  * Security Policy    -- one row per ACE, with source/dest/service resolved
  * Network Objects    -- name / type / value
  * Service Objects    -- name / protocol / ports
  * Network Groups     -- group / members / fully-resolved leaves
  * Service Groups     -- group / members / fully-resolved ports
  * Expanded Policy     -- (optional, --expand) ACEs exploded to the cartesian
                          product of resolved src x dst x service

Usage:
    asa_policy_to_xlsx.py running-config.txt
    asa_policy_to_xlsx.py running-config.txt -o policy.xlsx --expand

Best-effort parser: anything it can't interpret is preserved in the "Raw" column
so nothing is silently dropped.
"""
from __future__ import annotations

import argparse
import os
import re
import sys
from collections import OrderedDict

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError:  # pragma: no cover
    sys.exit("openpyxl is required: pip install openpyxl")

IPV4_RE = re.compile(r"^\d{1,3}(?:\.\d{1,3}){3}$")
PORT_OPS = ("eq", "neq", "lt", "gt")


# --------------------------------------------------------------------------- #
# Model
# --------------------------------------------------------------------------- #
class Config:
    """Symbol tables built from the config, plus resolution helpers."""

    def __init__(self):
        self.net_objects = OrderedDict()   # name -> {"kind","value","desc","raw"}
        self.svc_objects = OrderedDict()   # name -> {"protocol","detail","desc","raw"}
        self.net_groups = OrderedDict()    # name -> {"desc","members":[...]}
        self.svc_groups = OrderedDict()    # name -> {"protocol","desc","members":[...]}
        self.proto_groups = OrderedDict()  # name -> {"desc","members":[...]}
        self.icmp_groups = OrderedDict()   # name -> {"desc","members":[...]}
        self.acls = OrderedDict()          # name -> [ace,...]
        self.bindings = OrderedDict()      # acl_name -> "direction/interface"

    # ---- network resolution -------------------------------------------- #
    def resolve_net_object(self, name, seen=None):
        seen = seen or set()
        if name in seen:
            return [f"<loop:{name}>"]
        seen = seen | {name}
        obj = self.net_objects.get(name)
        if not obj:
            return [f"<undefined object {name}>"]
        return [obj["value"]]

    def resolve_net_group(self, name, seen=None):
        seen = seen or set()
        if name in seen:
            return [f"<loop:{name}>"]
        seen = seen | {name}
        grp = self.net_groups.get(name)
        if not grp:
            return [f"<undefined group {name}>"]
        out = []
        for m in grp["members"]:
            if m["type"] == "object":
                out += self.resolve_net_object(m["value"], seen)
            elif m["type"] == "group":
                out += self.resolve_net_group(m["value"], seen)
            else:
                out.append(m["value"])
        return out

    # ---- service resolution -------------------------------------------- #
    def resolve_svc_object(self, name, seen=None):
        seen = seen or set()
        if name in seen:
            return [f"<loop:{name}>"]
        obj = self.svc_objects.get(name)
        if not obj:
            return [f"<undefined service {name}>"]
        detail = obj["detail"]
        return [f"{obj['protocol']} {detail}".strip()]

    def resolve_svc_group(self, name, seen=None):
        seen = seen or set()
        if name in seen:
            return [f"<loop:{name}>"]
        seen = seen | {name}
        grp = self.svc_groups.get(name)
        if grp:
            proto = grp.get("protocol") or ""
            out = []
            for m in grp["members"]:
                if m["type"] == "object":
                    out += self.resolve_svc_object(m["value"], seen)
                elif m["type"] == "group":
                    out += self.resolve_svc_group(m["value"], seen)
                elif m["type"] == "port":
                    out.append(f"{proto}/{m['value']}".strip("/") if proto else m["value"])
                else:
                    out.append(m["value"])
            return out
        # protocol object-group used where a service group was expected
        if name in self.proto_groups:
            return self.resolve_proto_group(name, seen)
        return [f"<undefined service-group {name}>"]

    def resolve_proto_group(self, name, seen=None):
        seen = seen or set()
        if name in seen:
            return [f"<loop:{name}>"]
        seen = seen | {name}
        grp = self.proto_groups.get(name)
        if not grp:
            return [f"<undefined protocol-group {name}>"]
        out = []
        for m in grp["members"]:
            if m["type"] == "group":
                out += self.resolve_proto_group(m["value"], seen)
            else:
                out.append(m["value"])
        return out


# --------------------------------------------------------------------------- #
# Parsing
# --------------------------------------------------------------------------- #
def parse_service_body(tokens):
    """tokens after 'service'/'service-object' -> (protocol, detail_string)."""
    if not tokens:
        return ("", "")
    proto = tokens[0]
    rest = tokens[1:]
    parts = []
    i = 0
    while i < len(rest):
        t = rest[i]
        if t in ("source", "destination"):
            side = t
            op = rest[i + 1] if i + 1 < len(rest) else ""
            if op == "range":
                parts.append(f"{side} range {rest[i+2]}-{rest[i+3]}")
                i += 4
            elif op in PORT_OPS:
                parts.append(f"{side} {op} {rest[i+2]}")
                i += 3
            else:
                parts.append(f"{side} {op}")
                i += 2
        else:
            parts.append(t)
            i += 1
    return (proto, " ".join(parts))


def parse_config(text):
    cfg = Config()
    block = None          # ("kind", name)
    lines = text.splitlines()

    for raw_line in lines:
        line = raw_line.rstrip()
        if not line.strip() or line.lstrip().startswith("!"):
            block = None
            continue

        indented = line[0] in " \t"
        stripped = line.strip()
        tok = stripped.split()

        # ---- sub-command lines (belong to current block) --------------- #
        if indented and block:
            kind, name = block
            _parse_subline(cfg, kind, name, tok, stripped)
            continue

        block = None  # a non-indented line ends any block

        # ---- object network / service --------------------------------- #
        if tok[:2] == ["object", "network"]:
            name = tok[2]
            cfg.net_objects[name] = {"kind": "", "value": "", "desc": "", "raw": ""}
            block = ("net_object", name)
        elif tok[:2] == ["object", "service"]:
            name = tok[2]
            cfg.svc_objects[name] = {"protocol": "", "detail": "", "desc": "", "raw": ""}
            block = ("svc_object", name)

        # ---- object-group variants ------------------------------------ #
        elif tok[:2] == ["object-group", "network"]:
            name = tok[2]
            cfg.net_groups[name] = {"desc": "", "members": []}
            block = ("net_group", name)
        elif tok[:2] == ["object-group", "service"]:
            name = tok[2]
            proto = tok[3] if len(tok) > 3 else ""
            cfg.svc_groups[name] = {"protocol": proto, "desc": "", "members": []}
            block = ("svc_group", name)
        elif tok[:2] == ["object-group", "protocol"]:
            name = tok[2]
            cfg.proto_groups[name] = {"desc": "", "members": []}
            block = ("proto_group", name)
        elif tok[:2] == ["object-group", "icmp-type"]:
            name = tok[2]
            cfg.icmp_groups[name] = {"desc": "", "members": []}
            block = ("icmp_group", name)
        elif tok[0] == "object-group":
            # user/security/other group types we don't model -- skip cleanly
            block = ("ignore", tok[2] if len(tok) > 2 else "")

        # ---- access-list / access-group ------------------------------- #
        elif tok[0] == "access-list":
            _parse_access_list(cfg, tok, stripped)
        elif tok[0] == "access-group":
            # access-group NAME {in|out} interface IFACE
            if len(tok) >= 5 and tok[2] in ("in", "out") and tok[3] == "interface":
                cfg.bindings[tok[1]] = f"{tok[2]} / {tok[4]}"

    return cfg


def _parse_subline(cfg, kind, name, tok, stripped):
    if tok[0] == "description":
        desc = stripped[len("description"):].strip()
        target = {
            "net_object": cfg.net_objects, "svc_object": cfg.svc_objects,
            "net_group": cfg.net_groups, "svc_group": cfg.svc_groups,
            "proto_group": cfg.proto_groups, "icmp_group": cfg.icmp_groups,
        }.get(kind)
        if target and name in target:
            target[name]["desc"] = desc
        return

    if kind == "net_object":
        o = cfg.net_objects[name]
        o["raw"] = stripped
        if tok[0] == "host":
            o["kind"], o["value"] = "host", f"host {tok[1]}"
        elif tok[0] == "subnet":
            o["kind"], o["value"] = "subnet", " ".join(tok[1:])
        elif tok[0] == "range":
            o["kind"], o["value"] = "range", f"{tok[1]}-{tok[2]}"
        elif tok[0] == "fqdn":
            o["kind"], o["value"] = "fqdn", tok[-1]

    elif kind == "svc_object":
        if tok[0] == "service":
            proto, detail = parse_service_body(tok[1:])
            cfg.svc_objects[name].update(protocol=proto, detail=detail, raw=stripped)

    elif kind == "net_group":
        m = _parse_network_member(tok)
        if m:
            cfg.net_groups[name]["members"].append(m)

    elif kind == "svc_group":
        m = _parse_service_member(tok)
        if m:
            cfg.svc_groups[name]["members"].append(m)

    elif kind == "proto_group":
        if tok[0] == "protocol-object":
            cfg.proto_groups[name]["members"].append({"type": "proto", "value": tok[1]})
        elif tok[0] == "group-object":
            cfg.proto_groups[name]["members"].append({"type": "group", "value": tok[1]})

    elif kind == "icmp_group":
        if tok[0] == "icmp-object":
            cfg.icmp_groups[name]["members"].append({"type": "icmp", "value": tok[1]})
        elif tok[0] == "group-object":
            cfg.icmp_groups[name]["members"].append({"type": "group", "value": tok[1]})


def _parse_network_member(tok):
    if tok[0] == "network-object":
        if tok[1] == "host":
            return {"type": "host", "value": f"host {tok[2]}"}
        if tok[1] == "object":
            return {"type": "object", "value": tok[2]}
        return {"type": "subnet", "value": " ".join(tok[1:])}
    if tok[0] == "group-object":
        return {"type": "group", "value": tok[1]}
    return None


def _parse_service_member(tok):
    if tok[0] == "port-object":
        if tok[1] == "range":
            return {"type": "port", "value": f"{tok[2]}-{tok[3]}"}
        return {"type": "port", "value": " ".join(tok[1:])}
    if tok[0] == "service-object":
        if tok[1] == "object":
            return {"type": "object", "value": tok[2]}
        proto, detail = parse_service_body(tok[1:])
        return {"type": "service", "value": f"{proto} {detail}".strip()}
    if tok[0] == "group-object":
        return {"type": "group", "value": tok[1]}
    return None


def _parse_addr(tokens, i):
    """Return (display, ref_or_None, next_index). ref = (kind, name)."""
    t = tokens[i]
    if t in ("any", "any4", "any6"):
        return (t, None, i + 1)
    if t == "host":
        return (f"host {tokens[i+1]}", None, i + 2)
    if t == "interface":
        return (f"interface {tokens[i+1]}", None, i + 2)
    if t == "object":
        return (f"object {tokens[i+1]}", ("object", tokens[i + 1]), i + 2)
    if t == "object-group":
        return (f"group {tokens[i+1]}", ("object-group", tokens[i + 1]), i + 2)
    if ":" in t:  # bare IPv6 host or prefix
        return (t, None, i + 1)
    if IPV4_RE.match(t) and i + 1 < len(tokens) and IPV4_RE.match(tokens[i + 1]):
        return (f"{t} {tokens[i+1]}", None, i + 2)
    return (t, None, i + 1)


def _parse_port(tokens, i, svc_groups):
    if i >= len(tokens):
        return (None, i)
    t = tokens[i]
    if t in PORT_OPS:
        return (f"{t} {tokens[i+1]}", i + 2)
    if t == "range":
        return (f"range {tokens[i+1]}-{tokens[i+2]}", i + 3)
    if t == "object-group" and i + 1 < len(tokens) and tokens[i + 1] in svc_groups:
        return (f"group {tokens[i+1]}", i + 2)
    return (None, i)


def _parse_access_list(cfg, tok, stripped):
    name = tok[1]
    cfg.acls.setdefault(name, [])
    rest = tok[2:]
    i = 0
    if i < len(rest) and rest[i] == "line":
        i += 2

    if i < len(rest) and rest[i] == "remark":
        cfg.acls[name].append({"kind": "remark", "text": " ".join(rest[i + 1:]), "raw": stripped})
        return

    acl_type = ""
    if i < len(rest) and rest[i] in ("extended", "standard"):
        acl_type = rest[i]
        i += 1

    ace = {"kind": "ace", "type": acl_type, "raw": stripped,
           "action": "", "protocol": "", "service_ref": None,
           "src": "", "src_ref": None, "src_port": "",
           "dst": "", "dst_ref": None, "dst_port": "", "options": ""}
    try:
        if i < len(rest) and rest[i] in ("permit", "deny"):
            ace["action"] = rest[i]
            i += 1

        # protocol
        if rest[i] == "object":
            ace["protocol"] = f"object {rest[i+1]}"
            ace["service_ref"] = ("object", rest[i + 1])
            i += 2
        elif rest[i] == "object-group":
            ace["protocol"] = f"group {rest[i+1]}"
            ace["service_ref"] = ("object-group", rest[i + 1])
            i += 2
        else:
            ace["protocol"] = rest[i]
            i += 1

        if acl_type == "standard":
            # standard: only a destination network follows
            ace["dst"], ace["dst_ref"], i = _parse_addr(rest, i)
        else:
            ace["src"], ace["src_ref"], i = _parse_addr(rest, i)
            ace["src_port"], i = _parse_port(rest, i, cfg.svc_groups)
            ace["dst"], ace["dst_ref"], i = _parse_addr(rest, i)
            ace["dst_port"], i = _parse_port(rest, i, cfg.svc_groups)

        ace["options"] = " ".join(rest[i:])
    except IndexError:
        ace["options"] = (ace["options"] + " [parse-truncated]").strip()

    cfg.acls[name].append(ace)


# --------------------------------------------------------------------------- #
# Resolution helpers for report rows
# --------------------------------------------------------------------------- #
def _resolve_addr_ref(cfg, ref):
    if not ref:
        return ""
    kind, nm = ref
    leaves = cfg.resolve_net_object(nm) if kind == "object" else cfg.resolve_net_group(nm)
    return ", ".join(leaves)


def _resolve_service_ref(cfg, ref):
    if not ref:
        return ""
    kind, nm = ref
    if kind == "object":
        return ", ".join(cfg.resolve_svc_object(nm))
    if nm in cfg.svc_groups:
        return ", ".join(cfg.resolve_svc_group(nm))
    if nm in cfg.proto_groups:
        return ", ".join(cfg.resolve_proto_group(nm))
    return f"<undefined {nm}>"


# --------------------------------------------------------------------------- #
# XLSX output
# --------------------------------------------------------------------------- #
HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
TITLE_FONT = Font(bold=True, size=13, color="1F3864")
PERMIT_FILL = PatternFill("solid", fgColor="C6EFCE")
DENY_FILL = PatternFill("solid", fgColor="FFC7CE")
PERMIT_FONT = Font(color="006100", bold=True)
DENY_FONT = Font(color="9C0006", bold=True)
REMARK_FILL = PatternFill("solid", fgColor="FFF2CC")
STRIPE_FILL = PatternFill("solid", fgColor="F2F2F2")
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP_TOP = Alignment(vertical="top", wrap_text=True)
TOP = Alignment(vertical="top")


def _style_header(ws, ncols, row=1):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = BORDER


def _autosize(ws, widths):
    for idx, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = w


def build_workbook(cfg, expand=False):
    wb = Workbook()
    _sheet_summary(wb, cfg)
    _sheet_policy(wb, cfg)
    _sheet_net_objects(wb, cfg)
    _sheet_svc_objects(wb, cfg)
    _sheet_net_groups(wb, cfg)
    _sheet_svc_groups(wb, cfg)
    if expand:
        _sheet_expanded(wb, cfg)
    return wb


def _sheet_summary(wb, cfg):
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = "Cisco ASA Security Policy Report"
    ws["A1"].font = TITLE_FONT
    rows = [
        ("Network objects", len(cfg.net_objects)),
        ("Service objects", len(cfg.svc_objects)),
        ("Network object-groups", len(cfg.net_groups)),
        ("Service object-groups", len(cfg.svc_groups)),
        ("Protocol object-groups", len(cfg.proto_groups)),
        ("ICMP-type object-groups", len(cfg.icmp_groups)),
        ("Access-lists", len(cfg.acls)),
        ("Total ACEs", sum(len(a) for a in cfg.acls.values())),
    ]
    r = 3
    ws.cell(row=r, column=1, value="Counts").font = Font(bold=True)
    r += 1
    for label, n in rows:
        ws.cell(row=r, column=1, value=label)
        ws.cell(row=r, column=2, value=n)
        r += 1

    r += 1
    ws.cell(row=r, column=1, value="Interface bindings (access-group)").font = Font(bold=True)
    r += 1
    ws.cell(row=r, column=1, value="ACL").font = HEADER_FONT
    ws.cell(row=r, column=1).fill = HEADER_FILL
    ws.cell(row=r, column=2, value="Direction / Interface").font = HEADER_FONT
    ws.cell(row=r, column=2).fill = HEADER_FILL
    ws.cell(row=r, column=3, value="ACEs").font = HEADER_FONT
    ws.cell(row=r, column=3).fill = HEADER_FILL
    r += 1
    for acl in cfg.acls:
        ws.cell(row=r, column=1, value=acl)
        ws.cell(row=r, column=2, value=cfg.bindings.get(acl, "(not bound)"))
        ws.cell(row=r, column=3, value=len(cfg.acls[acl]))
        r += 1
    _autosize(ws, [30, 26, 10])


def _sheet_policy(wb, cfg):
    ws = wb.create_sheet("Security Policy")
    headers = ["ACL", "Applied", "#", "Action", "Protocol", "Source", "Src Port",
               "Destination", "Dst Port", "Service (resolved)", "Options",
               "Resolved Source", "Resolved Destination", "Remark", "Raw"]
    ws.append(headers)
    _style_header(ws, len(headers))

    row = 2
    for acl, aces in cfg.acls.items():
        pending_remark = ""
        seq = 0
        for ace in aces:
            if ace["kind"] == "remark":
                pending_remark = (pending_remark + " | " + ace["text"]).strip(" |") if pending_remark else ace["text"]
                # also emit a visible remark row for context
                ws.cell(row=row, column=1, value=acl)
                rc = ws.cell(row=row, column=14, value=ace["text"])
                for c in range(1, len(headers) + 1):
                    cell = ws.cell(row=row, column=c)
                    cell.fill = REMARK_FILL
                    cell.alignment = WRAP_TOP
                    cell.border = BORDER
                ws.cell(row=row, column=4, value="remark").font = Font(italic=True, color="7F6000")
                ws.cell(row=row, column=15, value=ace["raw"])
                row += 1
                continue

            seq += 1
            svc_res = _resolve_service_ref(cfg, ace["service_ref"])
            values = [
                acl,
                cfg.bindings.get(acl, ""),
                seq,
                ace["action"],
                ace["protocol"],
                ace["src"],
                ace["src_port"] or "",
                ace["dst"],
                ace["dst_port"] or "",
                svc_res,
                ace["options"],
                _resolve_addr_ref(cfg, ace["src_ref"]),
                _resolve_addr_ref(cfg, ace["dst_ref"]),
                pending_remark,
                ace["raw"],
            ]
            for c, v in enumerate(values, start=1):
                cell = ws.cell(row=row, column=c, value=v)
                cell.border = BORDER
                cell.alignment = WRAP_TOP if c in (10, 12, 13, 14, 15) else TOP
            # action colouring
            acell = ws.cell(row=row, column=4)
            if ace["action"] == "permit":
                acell.fill, acell.font = PERMIT_FILL, PERMIT_FONT
            elif ace["action"] == "deny":
                acell.fill, acell.font = DENY_FILL, DENY_FONT
            pending_remark = ""
            row += 1

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{max(row-1,1)}"
    _autosize(ws, [16, 16, 4, 8, 14, 22, 12, 22, 12, 30, 16, 30, 30, 26, 44])


def _sheet_net_objects(wb, cfg):
    ws = wb.create_sheet("Network Objects")
    headers = ["Name", "Type", "Value", "Description"]
    ws.append(headers)
    _style_header(ws, len(headers))
    r = 2
    for name, o in cfg.net_objects.items():
        for c, v in enumerate([name, o["kind"], o["value"], o["desc"]], start=1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.border = BORDER
            cell.alignment = TOP
        if r % 2 == 0:
            for c in range(1, len(headers) + 1):
                if not ws.cell(row=r, column=c).fill.fgColor.rgb or ws.cell(row=r, column=c).fill.patternType is None:
                    ws.cell(row=r, column=c).fill = STRIPE_FILL
        r += 1
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:D{max(r-1,1)}"
    _autosize(ws, [28, 10, 34, 40])


def _sheet_svc_objects(wb, cfg):
    ws = wb.create_sheet("Service Objects")
    headers = ["Name", "Protocol", "Ports / Detail", "Description"]
    ws.append(headers)
    _style_header(ws, len(headers))
    r = 2
    for name, o in cfg.svc_objects.items():
        for c, v in enumerate([name, o["protocol"], o["detail"], o["desc"]], start=1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.border = BORDER
            cell.alignment = TOP
        r += 1
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:D{max(r-1,1)}"
    _autosize(ws, [28, 10, 34, 40])


def _sheet_net_groups(wb, cfg):
    ws = wb.create_sheet("Network Groups")
    headers = ["Group", "Members (as configured)", "Resolved leaves", "Description"]
    ws.append(headers)
    _style_header(ws, len(headers))
    r = 2
    for name, g in cfg.net_groups.items():
        members = "\n".join(
            (f"object {m['value']}" if m["type"] == "object"
             else f"group {m['value']}" if m["type"] == "group"
             else m["value"])
            for m in g["members"]
        )
        resolved = "\n".join(cfg.resolve_net_group(name))
        for c, v in enumerate([name, members, resolved, g["desc"]], start=1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.border = BORDER
            cell.alignment = WRAP_TOP
        r += 1
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:D{max(r-1,1)}"
    _autosize(ws, [28, 34, 40, 30])


def _sheet_svc_groups(wb, cfg):
    ws = wb.create_sheet("Service Groups")
    headers = ["Group", "Protocol", "Members (as configured)", "Resolved ports/services", "Description"]
    ws.append(headers)
    _style_header(ws, len(headers))
    r = 2
    for name, g in cfg.svc_groups.items():
        members = "\n".join(
            (f"object {m['value']}" if m["type"] == "object"
             else f"group {m['value']}" if m["type"] == "group"
             else m["value"])
            for m in g["members"]
        )
        resolved = "\n".join(cfg.resolve_svc_group(name))
        for c, v in enumerate([name, g.get("protocol", ""), members, resolved, g["desc"]], start=1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.border = BORDER
            cell.alignment = WRAP_TOP
        r += 1
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:E{max(r-1,1)}"
    _autosize(ws, [28, 10, 34, 40, 30])


def _sheet_expanded(wb, cfg, cap=5000):
    """One row per (src, dst, service) leaf combination. Capped to avoid blow-up."""
    ws = wb.create_sheet("Expanded Policy")
    headers = ["ACL", "Applied", "Action", "Protocol", "Source", "Src Port",
               "Destination", "Dst Port", "Service"]
    ws.append(headers)
    _style_header(ws, len(headers))
    r = 2
    truncated = False
    for acl, aces in cfg.acls.items():
        for ace in aces:
            if ace["kind"] != "ace":
                continue
            srcs = _resolve_addr_ref(cfg, ace["src_ref"]).split(", ") if ace["src_ref"] else [ace["src"]]
            dsts = _resolve_addr_ref(cfg, ace["dst_ref"]).split(", ") if ace["dst_ref"] else [ace["dst"]]
            svcs = _resolve_service_ref(cfg, ace["service_ref"]).split(", ") if ace["service_ref"] else [ace["protocol"]]
            for s in srcs:
                for d in dsts:
                    for sv in svcs:
                        if r > cap + 1:
                            truncated = True
                            break
                        vals = [acl, cfg.bindings.get(acl, ""), ace["action"], ace["protocol"],
                                s, ace["src_port"] or "", d, ace["dst_port"] or "", sv]
                        for c, v in enumerate(vals, start=1):
                            cell = ws.cell(row=r, column=c, value=v)
                            cell.border = BORDER
                            cell.alignment = TOP
                        ac = ws.cell(row=r, column=3)
                        if ace["action"] == "permit":
                            ac.fill, ac.font = PERMIT_FILL, PERMIT_FONT
                        elif ace["action"] == "deny":
                            ac.fill, ac.font = DENY_FILL, DENY_FONT
                        r += 1
    if truncated:
        ws.cell(row=r, column=1, value=f"[truncated at {cap} rows -- use per-ACL views for full detail]")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:I{max(r-1,1)}"
    _autosize(ws, [16, 16, 8, 14, 24, 12, 24, 12, 24])


# --------------------------------------------------------------------------- #
def main(argv=None):
    ap = argparse.ArgumentParser(description="Render a Cisco ASA config as a readable XLSX policy report.")
    ap.add_argument("config", help="Path to the ASA running-config text file")
    ap.add_argument("-o", "--output", help="Output .xlsx path (default: <config>.xlsx)")
    ap.add_argument("--expand", action="store_true",
                    help="Add an 'Expanded Policy' sheet (ACEs exploded to resolved leaves)")
    args = ap.parse_args(argv)

    with open(args.config, "r", encoding="utf-8", errors="replace") as fh:
        text = fh.read()

    cfg = parse_config(text)
    wb = build_workbook(cfg, expand=args.expand)

    out = args.output or (os.path.splitext(args.config)[0] + ".xlsx")
    wb.save(out)

    print(f"Wrote {out}")
    print(f"  network objects : {len(cfg.net_objects)}")
    print(f"  service objects : {len(cfg.svc_objects)}")
    print(f"  network groups  : {len(cfg.net_groups)}")
    print(f"  service groups  : {len(cfg.svc_groups)}")
    print(f"  access-lists    : {len(cfg.acls)} ({sum(len(a) for a in cfg.acls.values())} lines)")
    return 0


if __name__ == "__main__":
    sys.exit(main())
